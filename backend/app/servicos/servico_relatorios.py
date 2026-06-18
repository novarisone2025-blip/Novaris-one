import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.modelos.produto import Produto


def formatar_moeda(valor: Decimal | float) -> str:
    texto = f"{float(valor):,.2f}"
    return "R$ " + texto.replace(",", "X").replace(".", ",").replace("X", ".")


def gerar_relatorio_pdf(
    produtos: list[Produto],
    nome_empresa: str,
) -> io.BytesIO:
    arquivo = io.BytesIO()
    documento = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"<b>Relatório de Estoque - {nome_empresa}</b>", estilos["Title"]),
        Spacer(1, 10),
    ]
    linhas = [[
        "Produto",
        "Código de barras",
        "Quantidade",
        "Valor unitário",
        "Valor total",
    ]]
    valor_geral = Decimal("0")
    for produto in produtos:
        valor_total = Decimal(produto.preco) * produto.quantidade
        valor_geral += valor_total
        linhas.append([
            produto.nome,
            produto.codigo_barras,
            str(produto.quantidade),
            formatar_moeda(produto.preco),
            formatar_moeda(valor_total),
        ])
    linhas.append(["", "", "", "Valor total geral", formatar_moeda(valor_geral)])

    tabela = Table(
        linhas,
        colWidths=[75 * mm, 48 * mm, 28 * mm, 36 * mm, 40 * mm],
    )
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6E5CF5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D7E3")),
        ("PADDING", (0, 0), (-1, -1), 7),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("FONTNAME", (3, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F0EDFF")),
    ]))
    elementos.append(tabela)
    documento.build(elementos)
    arquivo.seek(0)
    return arquivo


def gerar_relatorio_excel(
    produtos: list[Produto],
    nome_empresa: str,
) -> io.BytesIO:
    pasta = Workbook()
    planilha = pasta.active
    planilha.title = "Estoque"
    planilha.append([f"Relatório de Estoque - {nome_empresa}"])
    planilha.merge_cells("A1:E1")
    planilha["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    planilha["A1"].fill = PatternFill("solid", fgColor="6E5CF5")
    planilha["A1"].alignment = Alignment(horizontal="center")
    planilha.append([])
    planilha.append([
        "Produto",
        "Código de barras",
        "Quantidade",
        "Valor unitário",
        "Valor total",
    ])
    for celula in planilha[3]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="6E5CF5")

    valor_geral = Decimal("0")
    for produto in produtos:
        valor_total = Decimal(produto.preco) * produto.quantidade
        valor_geral += valor_total
        planilha.append([
            produto.nome,
            produto.codigo_barras,
            produto.quantidade,
            float(produto.preco),
            float(valor_total),
        ])
    planilha.append(["", "", "", "Valor total geral", float(valor_geral)])
    ultima_linha = planilha.max_row
    planilha.cell(ultima_linha, 4).font = Font(bold=True)
    planilha.cell(ultima_linha, 5).font = Font(bold=True)

    for linha in range(4, ultima_linha + 1):
        planilha.cell(linha, 4).number_format = 'R$ #,##0.00'
        planilha.cell(linha, 5).number_format = 'R$ #,##0.00'
    larguras = {"A": 36, "B": 24, "C": 14, "D": 18, "E": 18}
    for coluna, largura in larguras.items():
        planilha.column_dimensions[coluna].width = largura

    arquivo = io.BytesIO()
    pasta.save(arquivo)
    arquivo.seek(0)
    return arquivo

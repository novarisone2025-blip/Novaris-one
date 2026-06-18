import io
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.modelos.operacoes import DevolucaoVenda
from app.modelos.produto import Produto
from app.modelos.usuario import Usuario
from app.modelos.venda import Venda
from app.servicos.servico_relatorios import formatar_moeda


def resolver_periodo(
    periodo: str,
    data_inicial: date | None,
    data_final: date | None,
) -> tuple[date, date, datetime, datetime]:
    hoje = datetime.now(timezone.utc).date()
    if periodo == "dia":
        inicio = fim = data_inicial or hoje
    elif periodo == "semana":
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = hoje
    elif periodo == "personalizado":
        inicio = data_inicial or hoje
        fim = data_final or inicio
    else:
        inicio = hoje.replace(day=1)
        fim = hoje
    if fim < inicio:
        inicio, fim = fim, inicio
    return (
        inicio,
        fim,
        datetime.combine(inicio, time.min),
        datetime.combine(fim + timedelta(days=1), time.min),
    )


def gerar_relatorio_avancado(
    usuario: Usuario,
    sessao: Session,
    periodo: str = "mes",
    data_inicial: date | None = None,
    data_final: date | None = None,
    produto_id: int | None = None,
    categoria: str | None = None,
    usuario_id: int | None = None,
    caixa_id: int | None = None,
    forma_pagamento: str | None = None,
) -> dict:
    inicio, fim, inicio_dt, fim_dt = resolver_periodo(
        periodo,
        data_inicial,
        data_final,
    )
    consulta = (
        select(Venda)
        .options(
            selectinload(Venda.itens),
            selectinload(Venda.devolucoes).selectinload(DevolucaoVenda.itens),
        )
        .where(
            Venda.empresa_id == usuario.empresa_id,
            Venda.status == "pago",
            Venda.data_venda >= inicio_dt,
            Venda.data_venda < fim_dt,
        )
    )
    if usuario_id:
        consulta = consulta.where(Venda.usuario_id == usuario_id)
    if caixa_id:
        consulta = consulta.where(Venda.caixa_id == caixa_id)
    if forma_pagamento:
        consulta = consulta.where(Venda.forma_pagamento == forma_pagamento)
    vendas = sessao.scalars(consulta).unique().all()
    produtos = {
        produto.id: produto
        for produto in sessao.scalars(
            select(Produto).where(Produto.empresa_id == usuario.empresa_id)
        ).all()
    }

    por_produto = defaultdict(
        lambda: {
            "quantidade": 0,
            "faturamento": Decimal("0"),
            "custo": Decimal("0"),
        }
    )
    vendas_consideradas = set()

    def aceita(produto: Produto | None) -> bool:
        if not produto:
            return False
        if produto_id and produto.id != produto_id:
            return False
        if categoria and produto.categoria.lower() != categoria.lower():
            return False
        return True

    for venda in vendas:
        fator = (
            Decimal(venda.valor_total) / Decimal(venda.subtotal)
            if venda.subtotal
            else Decimal("0")
        )
        venda_tem_item = False
        for item in venda.itens:
            produto = produtos.get(item.produto_id)
            if not aceita(produto):
                continue
            venda_tem_item = True
            dados = por_produto[item.produto_id]
            dados["quantidade"] += item.quantidade
            dados["faturamento"] += (
                Decimal(item.valor_unitario) * item.quantidade * fator
            )
            dados["custo"] += Decimal(item.custo_total)
        for operacao in venda.devolucoes:
            if not (inicio_dt <= operacao.data_operacao < fim_dt):
                continue
            for item in operacao.itens:
                produto = produtos.get(item.produto_id)
                if not aceita(produto):
                    continue
                venda_tem_item = True
                dados = por_produto[item.produto_id]
                valor = Decimal(item.valor_unitario) * item.quantidade
                custo = Decimal(item.custo_unitario) * item.quantidade
                if item.direcao == "devolvido":
                    dados["quantidade"] -= item.quantidade
                    dados["faturamento"] -= valor * fator
                    dados["custo"] -= custo
                else:
                    dados["quantidade"] += item.quantidade
                    dados["faturamento"] += valor
                    dados["custo"] += custo
        if venda_tem_item:
            vendas_consideradas.add(venda.id)

    linhas_produtos = []
    for id_produto, dados in por_produto.items():
        produto = produtos[id_produto]
        linhas_produtos.append({
            "produto_id": produto.id,
            "nome": produto.nome,
            "codigo_barras": produto.codigo_barras,
            "categoria": produto.categoria,
            "quantidade": max(dados["quantidade"], 0),
            "faturamento": max(dados["faturamento"], Decimal("0")),
            "custo": max(dados["custo"], Decimal("0")),
        })
    linhas_produtos.sort(
        key=lambda item: (item["quantidade"], item["faturamento"]),
        reverse=True,
    )
    faturamento = sum(
        (item["faturamento"] for item in linhas_produtos),
        Decimal("0"),
    )
    custo = sum(
        (item["custo"] for item in linhas_produtos),
        Decimal("0"),
    )
    quantidade = sum(item["quantidade"] for item in linhas_produtos)
    quantidade_vendas = len(vendas_consideradas)
    return {
        "data_inicial": inicio,
        "data_final": fim,
        "faturamento": faturamento,
        "lucro": faturamento - custo,
        "quantidade_vendida": quantidade,
        "quantidade_vendas": quantidade_vendas,
        "ticket_medio": (
            faturamento / quantidade_vendas
            if quantidade_vendas
            else Decimal("0")
        ),
        "produtos_mais_vendidos": [
            {chave: valor for chave, valor in item.items() if chave != "custo"}
            for item in linhas_produtos[:20]
        ],
    }


def relatorio_avancado_pdf(
    dados: dict,
    nome_empresa: str,
) -> io.BytesIO:
    arquivo = io.BytesIO()
    documento = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(
            f"<b>Relatorio Avancado de Vendas - {nome_empresa}</b>",
            estilos["Title"],
        ),
        Paragraph(
            f"{dados['data_inicial'].strftime('%d/%m/%Y')} a "
            f"{dados['data_final'].strftime('%d/%m/%Y')}",
            estilos["Normal"],
        ),
        Spacer(1, 10),
        Paragraph(
            f"Faturamento: <b>{formatar_moeda(dados['faturamento'])}</b> | "
            f"Lucro: <b>{formatar_moeda(dados['lucro'])}</b> | "
            f"Unidades: <b>{dados['quantidade_vendida']}</b> | "
            f"Ticket medio: <b>{formatar_moeda(dados['ticket_medio'])}</b>",
            estilos["Normal"],
        ),
        Spacer(1, 12),
    ]
    linhas = [["Produto", "Codigo", "Categoria", "Quantidade", "Faturamento"]]
    for item in dados["produtos_mais_vendidos"]:
        linhas.append([
            item["nome"],
            item["codigo_barras"],
            item["categoria"],
            str(item["quantidade"]),
            formatar_moeda(item["faturamento"]),
        ])
    tabela = Table(
        linhas,
        colWidths=[75 * mm, 48 * mm, 50 * mm, 30 * mm, 38 * mm],
    )
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6E5CF5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D7E3")),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabela)
    documento.build(elementos)
    arquivo.seek(0)
    return arquivo


def relatorio_avancado_excel(
    dados: dict,
    nome_empresa: str,
) -> io.BytesIO:
    pasta = Workbook()
    planilha = pasta.active
    planilha.title = "Vendas"
    planilha.append([f"Relatorio Avancado de Vendas - {nome_empresa}"])
    planilha.append(["Periodo", dados["data_inicial"], dados["data_final"]])
    planilha.append(["Faturamento", float(dados["faturamento"])])
    planilha.append(["Lucro", float(dados["lucro"])])
    planilha.append(["Quantidade vendida", dados["quantidade_vendida"]])
    planilha.append(["Ticket medio", float(dados["ticket_medio"])])
    planilha.append([])
    planilha.append(["Produto", "Codigo", "Categoria", "Quantidade", "Faturamento"])
    for celula in planilha[8]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="6E5CF5")
    for item in dados["produtos_mais_vendidos"]:
        planilha.append([
            item["nome"],
            item["codigo_barras"],
            item["categoria"],
            item["quantidade"],
            float(item["faturamento"]),
        ])
    planilha["B3"].number_format = 'R$ #,##0.00'
    planilha["B4"].number_format = 'R$ #,##0.00'
    planilha["B6"].number_format = 'R$ #,##0.00'
    for linha in range(9, planilha.max_row + 1):
        planilha.cell(linha, 5).number_format = 'R$ #,##0.00'
    for coluna, largura in {
        "A": 38,
        "B": 24,
        "C": 24,
        "D": 16,
        "E": 20,
    }.items():
        planilha.column_dimensions[coluna].width = largura
    arquivo = io.BytesIO()
    pasta.save(arquivo)
    arquivo.seek(0)
    return arquivo

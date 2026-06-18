import io
from decimal import Decimal

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.esquemas.cliente import ClienteAtualizacao, ClienteCriacao
from app.modelos.cliente import Cliente
from app.modelos.usuario import Usuario
from app.modelos.venda import ItemVenda, Venda
from app.servicos.servico_auditoria import registrar_auditoria


def _texto_opcional(valor: str | None) -> str | None:
    texto = (valor or "").strip()
    return texto or None


def _validar_unicos(
    dados: ClienteCriacao | ClienteAtualizacao,
    usuario: Usuario,
    sessao: Session,
    cliente_id: int | None = None,
) -> None:
    filtros = [Cliente.empresa_id == usuario.empresa_id]
    alternativas = []
    email = _texto_opcional(str(dados.email) if dados.email else None)
    documento = _texto_opcional(dados.documento)
    if email:
        alternativas.append(Cliente.email == email.lower())
    if documento:
        alternativas.append(Cliente.documento == documento)
    if not alternativas:
        return
    consulta = select(Cliente).where(*filtros, or_(*alternativas))
    if cliente_id:
        consulta = consulta.where(Cliente.id != cliente_id)
    existente = sessao.scalar(consulta)
    if existente:
        raise HTTPException(
            409,
            "Ja existe um cliente com este e-mail ou CPF/CNPJ.",
        )


def salvar_cliente(
    dados: ClienteCriacao | ClienteAtualizacao,
    usuario: Usuario,
    sessao: Session,
    cliente_id: int | None = None,
) -> Cliente:
    _validar_unicos(dados, usuario, sessao, cliente_id)
    if cliente_id:
        cliente = buscar_cliente_modelo(cliente_id, usuario, sessao)
    else:
        cliente = Cliente(
            empresa_id=usuario.empresa_id,
            usuario_id=usuario.id,
            nome="",
        )
        sessao.add(cliente)
    cliente.nome = dados.nome.strip()
    cliente.documento = _texto_opcional(dados.documento)
    cliente.telefone = _texto_opcional(dados.telefone)
    cliente.whatsapp = _texto_opcional(dados.whatsapp)
    cliente.email = (
        str(dados.email).strip().lower() if dados.email else None
    )
    cliente.endereco = _texto_opcional(dados.endereco)
    cliente.observacoes = _texto_opcional(dados.observacoes)
    if isinstance(dados, ClienteAtualizacao):
        cliente.ativo = dados.ativo
    sessao.flush()
    registrar_auditoria(
        sessao,
        usuario,
        "cliente_atualizado" if cliente_id else "cliente_cadastrado",
        "cliente",
        cliente.id,
        {"nome": cliente.nome, "documento": cliente.documento},
    )
    sessao.commit()
    sessao.refresh(cliente)
    return cliente


def buscar_cliente_modelo(
    cliente_id: int,
    usuario: Usuario,
    sessao: Session,
) -> Cliente:
    cliente = sessao.scalar(
        select(Cliente).where(
            Cliente.id == cliente_id,
            Cliente.empresa_id == usuario.empresa_id,
        )
    )
    if not cliente:
        raise HTTPException(404, "Cliente nao encontrado.")
    return cliente


def _metricas_clientes(
    usuario: Usuario,
    sessao: Session,
) -> dict[int, dict]:
    linhas = sessao.execute(
        select(
            Venda.cliente_id,
            func.coalesce(func.sum(Venda.valor_total), 0),
            func.count(Venda.id),
            func.max(Venda.data_venda),
        ).where(
            Venda.empresa_id == usuario.empresa_id,
            Venda.status == "pago",
            Venda.cliente_id.is_not(None),
        ).group_by(Venda.cliente_id)
    ).all()
    return {
        cliente_id: {
            "total_gasto": Decimal(total),
            "quantidade_compras": quantidade,
            "ticket_medio": (
                Decimal(total) / quantidade if quantidade else Decimal("0")
            ),
            "ultima_compra": ultima,
        }
        for cliente_id, total, quantidade, ultima in linhas
    }


def serializar_cliente(cliente: Cliente, metricas: dict | None = None) -> dict:
    return {
        "id": cliente.id,
        "nome": cliente.nome,
        "documento": cliente.documento,
        "telefone": cliente.telefone,
        "whatsapp": cliente.whatsapp,
        "email": cliente.email,
        "endereco": cliente.endereco,
        "observacoes": cliente.observacoes,
        "ativo": cliente.ativo,
        "data_criacao": cliente.data_criacao,
        "data_atualizacao": cliente.data_atualizacao,
        **(metricas or {
            "total_gasto": Decimal("0"),
            "quantidade_compras": 0,
            "ticket_medio": Decimal("0"),
            "ultima_compra": None,
        }),
    }


def listar_clientes(
    usuario: Usuario,
    sessao: Session,
    busca: str = "",
    somente_ativos: bool = True,
) -> list[dict]:
    consulta = select(Cliente).where(Cliente.empresa_id == usuario.empresa_id)
    if somente_ativos:
        consulta = consulta.where(Cliente.ativo.is_(True))
    termo = busca.strip()
    if termo:
        consulta = consulta.where(or_(
            Cliente.nome.ilike(f"%{termo}%"),
            Cliente.documento.ilike(f"%{termo}%"),
            Cliente.telefone.ilike(f"%{termo}%"),
            Cliente.whatsapp.ilike(f"%{termo}%"),
            Cliente.email.ilike(f"%{termo}%"),
        ))
    clientes = sessao.scalars(consulta.order_by(Cliente.nome).limit(300)).all()
    metricas = _metricas_clientes(usuario, sessao)
    return [
        serializar_cliente(cliente, metricas.get(cliente.id))
        for cliente in clientes
    ]


def detalhar_cliente(
    cliente_id: int,
    usuario: Usuario,
    sessao: Session,
) -> dict:
    cliente = buscar_cliente_modelo(cliente_id, usuario, sessao)
    vendas = sessao.scalars(
        select(Venda).where(
            Venda.empresa_id == usuario.empresa_id,
            Venda.cliente_id == cliente.id,
            Venda.status == "pago",
        ).order_by(Venda.data_venda.desc())
    ).all()
    produtos = sessao.execute(
        select(
            ItemVenda.nome_produto,
            ItemVenda.codigo_barras,
            func.sum(ItemVenda.quantidade),
        )
        .join(Venda, Venda.id == ItemVenda.venda_id)
        .where(
            Venda.empresa_id == usuario.empresa_id,
            Venda.cliente_id == cliente.id,
            Venda.status == "pago",
        )
        .group_by(ItemVenda.nome_produto, ItemVenda.codigo_barras)
        .order_by(func.sum(ItemVenda.quantidade).desc())
        .limit(5)
    ).all()
    metricas = _metricas_clientes(usuario, sessao).get(cliente.id)
    return {
        **serializar_cliente(cliente, metricas),
        "historico_compras": [
            {
                "venda_id": venda.id,
                "data_venda": venda.data_venda,
                "valor_total": venda.valor_total,
                "forma_pagamento": venda.forma_pagamento,
                "quantidade_itens": sum(item.quantidade for item in venda.itens),
            }
            for venda in vendas
        ],
        "produtos_mais_comprados": [
            {
                "nome": nome,
                "codigo_barras": codigo,
                "quantidade": quantidade,
            }
            for nome, codigo, quantidade in produtos
        ],
    }


def gerar_relatorio_clientes_excel(clientes: list[dict], empresa: str) -> io.BytesIO:
    pasta = Workbook()
    planilha = pasta.active
    planilha.title = "Clientes"
    planilha.append([f"Relatorio de Clientes - {empresa}"])
    planilha.append([])
    planilha.append([
        "Nome", "CPF/CNPJ", "Telefone", "WhatsApp", "E-mail",
        "Total gasto", "Compras", "Ticket medio", "Ultima compra",
    ])
    for celula in planilha[3]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="6E5CF5")
    for cliente in clientes:
        planilha.append([
            cliente["nome"],
            cliente["documento"] or "",
            cliente["telefone"] or "",
            cliente["whatsapp"] or "",
            cliente["email"] or "",
            float(cliente["total_gasto"]),
            cliente["quantidade_compras"],
            float(cliente["ticket_medio"]),
            cliente["ultima_compra"],
        ])
    for linha in range(4, planilha.max_row + 1):
        planilha.cell(linha, 6).number_format = 'R$ #,##0.00'
        planilha.cell(linha, 8).number_format = 'R$ #,##0.00'
    for coluna, largura in zip("ABCDEFGHI", [28, 20, 18, 18, 28, 16, 12, 16, 20]):
        planilha.column_dimensions[coluna].width = largura
    arquivo = io.BytesIO()
    pasta.save(arquivo)
    arquivo.seek(0)
    return arquivo


def gerar_relatorio_clientes_pdf(clientes: list[dict], empresa: str) -> io.BytesIO:
    arquivo = io.BytesIO()
    documento = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    estilos = getSampleStyleSheet()
    linhas = [["Cliente", "Documento", "Contato", "Compras", "Total", "Ticket medio"]]
    for cliente in clientes:
        linhas.append([
            cliente["nome"],
            cliente["documento"] or "-",
            cliente["whatsapp"] or cliente["telefone"] or cliente["email"] or "-",
            str(cliente["quantidade_compras"]),
            f"R$ {Decimal(cliente['total_gasto']):,.2f}",
            f"R$ {Decimal(cliente['ticket_medio']):,.2f}",
        ])
    tabela = Table(linhas, colWidths=[55 * mm, 35 * mm, 55 * mm, 25 * mm, 30 * mm, 30 * mm])
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6E5CF5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#D9D7E3")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    documento.build([
        Paragraph(f"<b>Clientes - {empresa}</b>", estilos["Title"]),
        Spacer(1, 10),
        tabela,
    ])
    arquivo.seek(0)
    return arquivo

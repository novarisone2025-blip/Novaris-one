import io
import logging
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.esquemas.financeiro import LancamentoCriacao
from app.modelos.financeiro import LancamentoFinanceiro
from app.modelos.operacoes import (
    DevolucaoVenda,
    ItemDevolucaoVenda,
    MovimentacaoCaixa,
)
from app.modelos.usuario import Usuario
from app.modelos.venda import CancelamentoVenda, ItemVenda, Venda
from app.servicos.servico_relatorios import formatar_moeda
from app.servicos.servico_auditoria import registrar_auditoria


logger = logging.getLogger("novaris.financeiro")


def limites_periodo(
    data_inicial: date | None,
    data_final: date | None,
) -> tuple[date, date, datetime, datetime]:
    hoje = datetime.now(timezone.utc).date()
    inicio = data_inicial or hoje.replace(day=1)
    fim = data_final or hoje
    return (
        inicio,
        fim,
        datetime.combine(inicio, time.min),
        datetime.combine(fim + timedelta(days=1), time.min),
    )


def calcular_resumo_financeiro(
    usuario: Usuario,
    sessao: Session,
    data_inicial: date | None = None,
    data_final: date | None = None,
) -> dict:
    inicio, fim, inicio_dt, fim_dt = limites_periodo(data_inicial, data_final)
    faturamento, quantidade_vendas = sessao.execute(
        select(
            func.coalesce(func.sum(Venda.valor_total), 0),
            func.count(Venda.id),
        ).where(
            Venda.empresa_id == usuario.empresa_id,
            Venda.status == "pago",
            Venda.data_venda >= inicio_dt,
            Venda.data_venda < fim_dt,
        )
    ).one()
    custo_produtos = sessao.scalar(
        select(func.coalesce(func.sum(ItemVenda.custo_total), 0))
        .join(Venda, Venda.id == ItemVenda.venda_id)
        .where(
            Venda.empresa_id == usuario.empresa_id,
            Venda.status == "pago",
            Venda.data_venda >= inicio_dt,
            Venda.data_venda < fim_dt,
        )
    )
    outras_entradas = sessao.scalar(
        select(func.coalesce(func.sum(LancamentoFinanceiro.valor), 0)).where(
            LancamentoFinanceiro.empresa_id == usuario.empresa_id,
            LancamentoFinanceiro.tipo == "entrada",
            LancamentoFinanceiro.data_lancamento >= inicio_dt,
            LancamentoFinanceiro.data_lancamento < fim_dt,
        )
    )
    despesas = sessao.scalar(
        select(func.coalesce(func.sum(LancamentoFinanceiro.valor), 0)).where(
            LancamentoFinanceiro.empresa_id == usuario.empresa_id,
            LancamentoFinanceiro.tipo == "saida",
            LancamentoFinanceiro.data_lancamento >= inicio_dt,
            LancamentoFinanceiro.data_lancamento < fim_dt,
        )
    )
    ajustes_devolucao = sessao.execute(
        select(
            func.coalesce(func.sum(DevolucaoVenda.valor_adicional), 0),
            func.coalesce(func.sum(DevolucaoVenda.valor_estornado), 0),
        ).where(
            DevolucaoVenda.empresa_id == usuario.empresa_id,
            DevolucaoVenda.data_operacao >= inicio_dt,
            DevolucaoVenda.data_operacao < fim_dt,
        )
    ).one()
    custos_devolucao = sessao.execute(
        select(
            func.coalesce(
                func.sum(
                    ItemDevolucaoVenda.custo_unitario
                    * ItemDevolucaoVenda.quantidade
                ).filter(ItemDevolucaoVenda.direcao == "devolvido"),
                0,
            ),
            func.coalesce(
                func.sum(
                    ItemDevolucaoVenda.custo_unitario
                    * ItemDevolucaoVenda.quantidade
                ).filter(ItemDevolucaoVenda.direcao == "novo"),
                0,
            ),
        )
        .join(
            DevolucaoVenda,
            DevolucaoVenda.id == ItemDevolucaoVenda.devolucao_id,
        )
        .where(
            DevolucaoVenda.empresa_id == usuario.empresa_id,
            DevolucaoVenda.data_operacao >= inicio_dt,
            DevolucaoVenda.data_operacao < fim_dt,
        )
    ).one()
    movimentos_caixa = sessao.execute(
        select(
            func.coalesce(
                func.sum(MovimentacaoCaixa.valor).filter(
                    MovimentacaoCaixa.tipo == "reforco"
                ),
                0,
            ),
            func.coalesce(
                func.sum(MovimentacaoCaixa.valor).filter(
                    MovimentacaoCaixa.tipo == "sangria"
                ),
                0,
            ),
        ).where(
            MovimentacaoCaixa.empresa_id == usuario.empresa_id,
            MovimentacaoCaixa.data_movimento >= inicio_dt,
            MovimentacaoCaixa.data_movimento < fim_dt,
        )
    ).one()
    faturamento = (
        Decimal(faturamento)
        + Decimal(ajustes_devolucao[0])
        - Decimal(ajustes_devolucao[1])
    )
    custo_produtos = (
        Decimal(custo_produtos)
        - Decimal(custos_devolucao[0])
        + Decimal(custos_devolucao[1])
    )
    outras_entradas = Decimal(outras_entradas)
    despesas = Decimal(despesas)
    lucro_bruto = faturamento - custo_produtos
    saldo_caixa = (
        faturamento
        + outras_entradas
        + Decimal(movimentos_caixa[0])
        - despesas
        - Decimal(movimentos_caixa[1])
    )
    return {
        "faturamento": faturamento,
        "custo_produtos": custo_produtos,
        "lucro_bruto": lucro_bruto,
        "outras_entradas": outras_entradas,
        "despesas": despesas,
        "saldo_caixa": saldo_caixa,
        "margem_bruta": (
            round(float(lucro_bruto / faturamento * 100), 2)
            if faturamento
            else 0
        ),
        "quantidade_vendas": quantidade_vendas,
        "ticket_medio": (
            faturamento / quantidade_vendas
            if quantidade_vendas
            else Decimal("0")
        ),
        "data_inicial": inicio,
        "data_final": fim,
    }


def criar_lancamento(
    dados: LancamentoCriacao,
    usuario: Usuario,
    sessao: Session,
) -> LancamentoFinanceiro:
    lancamento = LancamentoFinanceiro(
        empresa_id=usuario.empresa_id,
        usuario_id=usuario.id,
        tipo=dados.tipo,
        categoria=dados.categoria.strip(),
        descricao=dados.descricao.strip(),
        valor=dados.valor,
        data_lancamento=(
            dados.data_lancamento
            or datetime.now(timezone.utc).replace(tzinfo=None)
        ),
    )
    sessao.add(lancamento)
    sessao.flush()
    registrar_auditoria(
        sessao,
        usuario,
        "lancamento_financeiro_criado",
        "lancamento_financeiro",
        lancamento.id,
        {
            "tipo": lancamento.tipo,
            "categoria": lancamento.categoria,
            "valor": lancamento.valor,
        },
    )
    sessao.commit()
    sessao.refresh(lancamento)
    logger.info(
        "lancamento financeiro registrado empresa_id=%s usuario_id=%s "
        "lancamento_id=%s tipo=%s valor=%s",
        usuario.empresa_id,
        usuario.id,
        lancamento.id,
        lancamento.tipo,
        lancamento.valor,
    )
    return lancamento


def listar_fluxo_caixa(
    usuario: Usuario,
    sessao: Session,
    data_inicial: date | None = None,
    data_final: date | None = None,
) -> list[dict]:
    _, _, inicio_dt, fim_dt = limites_periodo(data_inicial, data_final)
    vendas = sessao.scalars(
        select(Venda)
        .options(selectinload(Venda.cancelamento))
        .where(
            Venda.empresa_id == usuario.empresa_id,
            or_(
                and_(
                    Venda.data_venda >= inicio_dt,
                    Venda.data_venda < fim_dt,
                ),
                Venda.cancelamento.has(
                    and_(
                        CancelamentoVenda.data_cancelamento >= inicio_dt,
                        CancelamentoVenda.data_cancelamento < fim_dt,
                    )
                ),
            ),
            Venda.status.in_(["pago", "cancelado"]),
        )
    ).all()
    lancamentos = sessao.scalars(
        select(LancamentoFinanceiro).where(
            LancamentoFinanceiro.empresa_id == usuario.empresa_id,
            LancamentoFinanceiro.data_lancamento >= inicio_dt,
            LancamentoFinanceiro.data_lancamento < fim_dt,
        )
    ).all()
    devolucoes = sessao.scalars(
        select(DevolucaoVenda).where(
            DevolucaoVenda.empresa_id == usuario.empresa_id,
            DevolucaoVenda.data_operacao >= inicio_dt,
            DevolucaoVenda.data_operacao < fim_dt,
        )
    ).all()
    movimentos_caixa = sessao.scalars(
        select(MovimentacaoCaixa).where(
            MovimentacaoCaixa.empresa_id == usuario.empresa_id,
            MovimentacaoCaixa.data_movimento >= inicio_dt,
            MovimentacaoCaixa.data_movimento < fim_dt,
        )
    ).all()
    autores = {
        autor.id: autor
        for autor in sessao.scalars(
            select(Usuario).where(Usuario.empresa_id == usuario.empresa_id)
        ).all()
    }
    fluxo = []
    for venda in vendas:
        if inicio_dt <= venda.data_venda < fim_dt:
            fluxo.append({
                "id": f"venda-{venda.id}",
                "tipo": "entrada",
                "categoria": "Vendas",
                "descricao": f"Venda #{venda.id}",
                "valor": venda.valor_total,
                "data_lancamento": venda.data_venda,
                "origem": "venda",
                "usuario_id": venda.usuario_id,
                "nome_usuario": autores[venda.usuario_id].nome,
                "cargo_usuario": autores[venda.usuario_id].cargo,
                "caixa_id": venda.caixa_id,
                "forma_pagamento": venda.forma_pagamento,
                "valor_recebido": venda.valor_recebido,
                "troco_entregue": venda.troco_entregue,
            })
        cancelamento = venda.cancelamento
        if (
            cancelamento
            and inicio_dt <= cancelamento.data_cancelamento < fim_dt
        ):
            autor = autores[cancelamento.usuario_id]
            fluxo.append({
                "id": f"estorno-{venda.id}",
                "tipo": "saida",
                "categoria": "Estornos",
                "descricao": f"Estorno da venda #{venda.id}",
                "valor": venda.valor_total,
                "data_lancamento": cancelamento.data_cancelamento,
                "origem": "estorno",
                "usuario_id": cancelamento.usuario_id,
                "nome_usuario": autor.nome,
                "cargo_usuario": autor.cargo,
                "caixa_id": venda.caixa_id,
            })
    fluxo.extend(
        {
            "id": lancamento.id,
            "tipo": lancamento.tipo,
            "categoria": lancamento.categoria,
            "descricao": lancamento.descricao,
            "valor": lancamento.valor,
            "data_lancamento": lancamento.data_lancamento,
            "origem": "manual",
            "usuario_id": lancamento.usuario_id,
            "nome_usuario": autores[lancamento.usuario_id].nome,
            "cargo_usuario": autores[lancamento.usuario_id].cargo,
            "caixa_id": None,
        }
        for lancamento in lancamentos
    )
    for operacao in devolucoes:
        autor = autores[operacao.usuario_id]
        if operacao.valor_estornado:
            fluxo.append({
                "id": f"devolucao-{operacao.id}",
                "tipo": "saida",
                "categoria": "Devolucoes",
                "descricao": (
                    f"{operacao.tipo.title()} da venda #{operacao.venda_id}"
                ),
                "valor": operacao.valor_estornado,
                "data_lancamento": operacao.data_operacao,
                "origem": "devolucao",
                "usuario_id": operacao.usuario_id,
                "nome_usuario": autor.nome,
                "cargo_usuario": autor.cargo,
                "caixa_id": operacao.caixa_id,
            })
        if operacao.valor_adicional:
            fluxo.append({
                "id": f"troca-adicional-{operacao.id}",
                "tipo": "entrada",
                "categoria": "Trocas",
                "descricao": f"Adicional da troca da venda #{operacao.venda_id}",
                "valor": operacao.valor_adicional,
                "data_lancamento": operacao.data_operacao,
                "origem": "troca",
                "usuario_id": operacao.usuario_id,
                "nome_usuario": autor.nome,
                "cargo_usuario": autor.cargo,
                "caixa_id": operacao.caixa_id,
            })
    for movimento in movimentos_caixa:
        autor = autores[movimento.usuario_id]
        fluxo.append({
            "id": f"caixa-{movimento.id}",
            "tipo": "saida" if movimento.tipo == "sangria" else "entrada",
            "categoria": movimento.tipo.title(),
            "descricao": movimento.motivo,
            "valor": movimento.valor,
            "data_lancamento": movimento.data_movimento,
            "origem": "caixa",
            "usuario_id": movimento.usuario_id,
            "nome_usuario": autor.nome,
            "cargo_usuario": autor.cargo,
            "caixa_id": movimento.caixa_id,
        })
    return sorted(
        fluxo,
        key=lambda item: item["data_lancamento"],
        reverse=True,
    )


def gerar_relatorio_financeiro_pdf(
    resumo: dict,
    fluxo: list[dict],
    nome_empresa: str,
) -> io.BytesIO:
    arquivo = io.BytesIO()
    documento = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        rightMargin=13 * mm,
        leftMargin=13 * mm,
        topMargin=13 * mm,
        bottomMargin=13 * mm,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"<b>Relatorio Financeiro - {nome_empresa}</b>", estilos["Title"]),
        Paragraph(
            f"{resumo['data_inicial'].strftime('%d/%m/%Y')} a "
            f"{resumo['data_final'].strftime('%d/%m/%Y')}",
            estilos["Normal"],
        ),
        Spacer(1, 10),
        Paragraph(
            f"Faturamento: <b>{formatar_moeda(resumo['faturamento'])}</b> | "
            f"Lucro bruto: <b>{formatar_moeda(resumo['lucro_bruto'])}</b> | "
            f"Saldo: <b>{formatar_moeda(resumo['saldo_caixa'])}</b>",
            estilos["Normal"],
        ),
        Spacer(1, 12),
    ]
    linhas = [[
        "Data",
        "Tipo",
        "Categoria",
        "Descricao",
        "Operador / Caixa",
        "Recebido",
        "Troco",
        "Valor",
    ]]
    for item in fluxo:
        linhas.append([
            item["data_lancamento"].strftime("%d/%m/%Y %H:%M"),
            item["tipo"].title(),
            item["categoria"],
            item["descricao"],
            (
                f"{item['nome_usuario']} / Caixa #{item['caixa_id']}"
                if item["caixa_id"]
                else item["nome_usuario"]
            ),
            (
                formatar_moeda(item["valor_recebido"])
                if item.get("valor_recebido") is not None
                else "-"
            ),
            (
                formatar_moeda(item["troco_entregue"])
                if item.get("troco_entregue") is not None
                else "-"
            ),
            formatar_moeda(item["valor"]),
        ])
    tabela = Table(
        linhas,
        colWidths=[
            28 * mm,
            18 * mm,
            25 * mm,
            58 * mm,
            43 * mm,
            25 * mm,
            23 * mm,
            25 * mm,
        ],
    )
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6E5CF5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D7E3")),
        ("PADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabela)
    documento.build(elementos)
    arquivo.seek(0)
    return arquivo


def gerar_relatorio_financeiro_excel(
    resumo: dict,
    fluxo: list[dict],
    nome_empresa: str,
) -> io.BytesIO:
    pasta = Workbook()
    planilha = pasta.active
    planilha.title = "Financeiro"
    planilha.append([f"Relatorio Financeiro - {nome_empresa}"])
    planilha.append(["Faturamento", float(resumo["faturamento"])])
    planilha.append(["Custo dos produtos", float(resumo["custo_produtos"])])
    planilha.append(["Lucro bruto", float(resumo["lucro_bruto"])])
    planilha.append(["Despesas", float(resumo["despesas"])])
    planilha.append(["Saldo de caixa", float(resumo["saldo_caixa"])])
    planilha.append([])
    planilha.append([
        "Data",
        "Tipo",
        "Categoria",
        "Descricao",
        "Operador",
        "Caixa",
        "Forma de pagamento",
        "Valor recebido",
        "Troco entregue",
        "Valor",
    ])
    for celula in planilha[8]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="6E5CF5")
    for item in fluxo:
        planilha.append([
            item["data_lancamento"],
            item["tipo"].title(),
            item["categoria"],
            item["descricao"],
            item["nome_usuario"],
            f"Caixa #{item['caixa_id']}" if item["caixa_id"] else "-",
            item.get("forma_pagamento") or "-",
            (
                float(item["valor_recebido"])
                if item.get("valor_recebido") is not None
                else None
            ),
            (
                float(item["troco_entregue"])
                if item.get("troco_entregue") is not None
                else None
            ),
            float(item["valor"]),
        ])
    for linha in range(2, 7):
        planilha.cell(linha, 2).number_format = 'R$ #,##0.00'
    for linha in range(9, planilha.max_row + 1):
        for coluna in (8, 9, 10):
            planilha.cell(linha, coluna).number_format = 'R$ #,##0.00'
    for coluna, largura in {
        "A": 22,
        "B": 18,
        "C": 24,
        "D": 40,
        "E": 24,
        "F": 14,
        "G": 20,
        "H": 18,
        "I": 18,
        "J": 18,
    }.items():
        planilha.column_dimensions[coluna].width = largura
    arquivo = io.BytesIO()
    pasta.save(arquivo)
    arquivo.seek(0)
    return arquivo
